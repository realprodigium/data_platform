import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

YEAR = 2025

NOMBRES_EMPLEADOS = [
    'ALFONSO CORREA', 'DAGOBERTO MUÑOZ', 'DANIEL TORO', 'DELIMIRO CASTRO', 'FABIO MARIN',
    'FELIX GUTIERREZ', 'FRANCISCO SIERRA', 'HANNER SANABRIA', 'HUGO HERNAN VARGAS',
    'JAIDER REYES', 'JEIFER PARRA', 'JESUS ENOC ANGARITA', 'JORGE ENRIQUE CASTRO',
    'JOSE ACOSTA', 'JOSE ARIZA OTERO', 'JUAN CAMILO GONZALEZ', 'JUAN SEBASTIAN MUÑOZ',
    'KEVIN BARBOSA', 'KEVIN GARCIA', 'KEVIN STIVEN BASTIDAS', 'LEWIS ZABALETA',
    'LUIS FERNANDEZ', 'LUIS GABRIEL GELVEZ', 'LUIS MIGUEL FERNANDEZ', 'LUIS TERNERA',
    'LUIS VIZCAINO', 'LUZ GOMEZ ARROYO', 'MARBIN ZABALETA', 'MARIO PERTUZ',
    'MIGUEL CASTRILLON', 'OVELIA INES VARGAS', 'PEDRO ORTIZ MERCADO', 'RAFAEL ARIZA',
    'ROLANDO VALDERRAMA', 'SAMUEL CARRILLO', 'TONYS VILLALOBOS', 'VICTOR MARQUEZ',
    'YEISON RAMIREZ'
]

NOMBRES_PROVEEDORES = [
    'INVERSIONES ZONA B', 'CASATORO SA BIC', 'CHANEME COMERCI', 'CLOUDFLEET SAS',
    'COMERCIALIZADORA', 'DISTRITODO VS', 'ERICK SANCHEZ A', 'EUGENIO CESAR M',
    'FREDERIC SANCHEZ', 'GRUPO EDS AUTOG', 'INVERSIONES COS', 'INVERSIONES GAL',
    'INVERSIONES URB', 'Ivon Lizarazo', 'JHORMAN DE JESUS', 'JOAN CARDONA QU',
    'JOSE LUIS ACOSTA', 'KELLY JOHANA SERNA', 'LA PERLA INVERS', 'LAURA VANESSA G',
    'LEONARDO MENDOZA', 'LETICIA ROMERO', 'luis padilla le', 'MANUEL YANES',
    'MICHAEL LARA OL', 'REPUEXCAVADORAS', 'RUBEN DARIO SUAREZ', 'SUMINISTROS IND',
    'TALLERES AUTORI', 'TECNODIESEL SAS', 'WALTER ANDRES G', 'WILLIAM BARROS',
    'YAMID ARTETA FA', 'YASSER BANDERAS', 'YENNIFER OSPINO', 'SERVIAGREGADOS'
]

# Definición de tipos de transacción
TRANSACCIONES = {
    'ABONO INTERESES AHORROS': {
        'credito': True,
        'rango': (10, 2000),  # intereses pequeños
        'prob': 0.12,
        'desc': 'ABONO INTERESES AHORROS'
    },
    'IMPTO GOBIERNO 4X1000': {
        'credito': False,
        'rango': (5000, 300000),  # impuesto 4x1000
        'prob': 0.10,
        'desc': 'IMPTO GOBIERNO 4X1000'
    },
    'PAGO A NOMIN': {
        'credito': False,
        'rango': (800000, 3500000),  # nómina realista
        'prob': 0.18,
        'desc': lambda: f'PAGO A NOMIN {random.choice(NOMBRES_EMPLEADOS)}'
    },
    'PAGO A PROVE': {
        'credito': False,
        'rango': (500000, 8000000),  # pagos a proveedores
        'prob': 0.15,
        'desc': lambda: f'PAGO A PROVE {random.choice(NOMBRES_PROVEEDORES)}'
    },
    'CONSIGNACION CORRESPONSAL CB': {
        'credito': True,
        'rango': (500000, 5000000),
        'prob': 0.05,
        'desc': 'CONSIGNACION CORRESPONSAL CB'
    },
    'CUOTA PLAN CANAL NEGOCIOS': {
        'credito': False,
        'rango': (30000, 200000),
        'prob': 0.03,
        'desc': 'CUOTA PLAN CANAL NEGOCIOS'
    },
    'COBRO IVA PAGOS AUTOMATICOS': {
        'credito': False,
        'rango': (500, 2000),
        'prob': 0.02,
        'desc': 'COBRO IVA PAGOS AUTOMATICOS'
    },
    'IVA CUOTA PLAN CANAL NEGOCIOS': {
        'credito': False,
        'rango': (5000, 50000),
        'prob': 0.02,
        'desc': 'IVA CUOTA PLAN CANAL NEGOCIOS'
    },
    'Pago cuota operacion Leasing': {
        'credito': False,
        'rango': (5000000, 15000000),
        'prob': 0.03,
        'desc': 'Pago cuota operacion Leasing'
    },
    'PAGO PSE': {
        'credito': False,
        'rango': (1000000, 20000000),
        'prob': 0.08,
        'desc': lambda: f'PAGO PSE {random.choice(["APORTES EN LINEA", "ASOPAGOS", "Hughes De Colombia S", "PATRIMONIOS AUTONOMO"])}'
    },
    'PAGO INTERBANC': {
        'credito': True,
        'rango': (10000000, 80000000),
        'prob': 0.04,
        'desc': lambda: f'PAGO INTERBANC {random.choice(["CONSORCIO EDUBA", "CONSORCIO HIDRO"])}'
    },
    'SERVICIO PAGO A PROVEEDORES': {
        'credito': False,
        'rango': (3000, 5000),
        'prob': 0.03,
        'desc': 'SERVICIO PAGO A PROVEEDORES'
    },
    'SERVICIO PAGO DE NOMINA': {
        'credito': False,
        'rango': (3000, 5000),
        'prob': 0.03,
        'desc': 'SERVICIO PAGO DE NOMINA'
    },
    'SERVICIO POR PAGOS A NEQUI': {
        'credito': False,
        'rango': (3000, 5000),
        'prob': 0.02,
        'desc': 'SERVICIO POR PAGOS A NEQUI'
    },
    'SERVICIO PAGO A OTROS BANCOS': {
        'credito': False,
        'rango': (5000, 10000),
        'prob': 0.02,
        'desc': 'SERVICIO PAGO A OTROS BANCOS'
    },
    'PAGO DE PROV': {
        'credito': True,
        'rango': (1000000, 25000000),
        'prob': 0.08,
        'desc': lambda: f'PAGO DE PROV {random.choice(NOMBRES_PROVEEDORES)}'
    }
}

def generar_fechas_trimestre(trimestre):
    """Devuelve lista de fechas (datetime) para el trimestre dado (1-4)"""
    if trimestre == 1:
        meses = [1, 2, 3]
    elif trimestre == 2:
        meses = [4, 5, 6]
    elif trimestre == 3:
        meses = [7, 8, 9]
    else:
        meses = [10, 11, 12]

    fechas = []
    for mes in meses:
        dias = list(range(1, 29))  # hasta 28 para evitar problemas de fin de mes
        for dia in dias:
            fechas.append(datetime(YEAR, mes, dia))
    return sorted(fechas)

def generar_saldo_inicial(trimestre):
    """Saldo inicial coherente para cada trimestre (en pesos colombianos)"""
    # Q1: bajo, Q2: crece, Q3: más alto, Q4: alto
    saldos = {1: 50_000_000, 2: 120_000_000, 3: 200_000_000, 4: 280_000_000}
    return saldos.get(trimestre, 100_000_000)

def generar_transacciones_trimestre(trimestre, num_transacciones_min=150, num_transacciones_max=220):
    """Genera lista de transacciones para un trimestre con saldos coherentes"""
    fechas = generar_fechas_trimestre(trimestre)
    saldo = generar_saldo_inicial(trimestre)
    transacciones = []

    # Determinar cuántas transacciones totales
    total_transacciones = random.randint(num_transacciones_min, num_transacciones_max)

    # Distribuir transacciones entre las fechas
    trans_por_dia = {}
    for _ in range(total_transacciones):
        fecha = random.choice(fechas)
        trans_por_dia.setdefault(fecha, 0)
        trans_por_dia[fecha] += 1

    # Generar transacciones por día
    tipos = list(TRANSACCIONES.keys())
    probs = [TRANSACCIONES[t]['prob'] for t in tipos]

    for fecha, cantidad in sorted(trans_por_dia.items()):
        for _ in range(cantidad):
            # Elegir tipo de transacción
            tipo = np.random.choice(tipos, p=probs)
            config = TRANSACCIONES[tipo]

            # Generar monto
            monto = random.randint(config['rango'][0], config['rango'][1])
            if not config['credito']:
                monto = -monto

            # Actualizar saldo
            saldo += monto

            # Descripción
            if callable(config['desc']):
                desc = config['desc']()
            else:
                desc = config['desc']

            # Sucursal (opcional, solo para algunas)
            sucursal = ''
            if tipo == 'CONSIGNACION CORRESPONSAL CB' and random.random() > 0.5:
                sucursal = 'CANAL CORRESPONSA'

            transacciones.append({
                'FECHA': fecha,
                'DESCRIPCIÓN': desc,
                'SUCURSAL': sucursal,
                'DCTO.': '',  # vacío
                'VALOR': monto,
                'SALDO': saldo,
                'Column1': '',
                '_1': ''
            })

    # Ordenar por fecha
    transacciones.sort(key=lambda x: x['FECHA'])

    # Agregar una fila final de "FIN ESTADO DE CUENTA"
    transacciones.append({
        'FECHA': None,
        'DESCRIPCIÓN': 'FIN ESTADO DE CUENTA',
        'SUCURSAL': '',
        'DCTO.': '',
        'VALOR': '',
        'SALDO': '',
        'Column1': '',
        '_1': ''
    })

    return transacciones

def guardar_excel(transacciones, nombre_archivo):
    """Guarda las transacciones en un archivo Excel con formato"""
    df = pd.DataFrame(transacciones)
    # Asegurar orden de columnas
    columnas = ['FECHA', 'DESCRIPCIÓN', 'SUCURSAL', 'DCTO.', 'VALOR', 'SALDO', 'Column1', '_1']
    df = df[columnas]

    # Formato de fechas
    df['FECHA'] = pd.to_datetime(df['FECHA']).dt.strftime('%Y-%m-%d %H:%M:%S')
    df['FECHA'] = df['FECHA'].replace('NaT', '')

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hoja2', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Hoja2']

        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # Formato de números (alineación derecha para VALOR y SALDO)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in (5, 6):  # VALOR y SALDO
                    cell.alignment = Alignment(horizontal='right')
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

def main():
    """Genera los cuatro archivos trimestrales"""
    print("Generando archivos trimestrales...")
    for trimestre in range(1, 5):
        nombre = f'transacciones_Q{trimestre}_{YEAR}.xlsx'
        print(f"Generando {nombre}...")
        transacciones = generar_transacciones_trimestre(trimestre)
        guardar_excel(transacciones, nombre)
        print(f"  → {len(transacciones)-1} transacciones generadas.")
    print("✅ Proceso completado. Los archivos están en el directorio actual.")

if __name__ == '__main__':
    main()